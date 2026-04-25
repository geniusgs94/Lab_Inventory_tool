import io
import re

import openpyxl
from fastapi import APIRouter, File, Request, UploadFile
from fastapi.responses import JSONResponse, RedirectResponse, StreamingResponse
from psycopg2.extras import RealDictCursor

from dependencies import _require_login, flash
from services.db import get_db_connection, return_db_connection

router = APIRouter()

EXPORT_COLUMNS = [
    "mac_address", "device_model", "owner", "availability",
    "reporting_manager", "team", "ip_address", "location", "lease",
    "project", "console", "power",
]
REQUIRED_COLUMNS = {"mac_address", "device_model", "availability"}
VALID_AVAILABILITY = {"Available", "In Use"}
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _require_admin(request):
    user = _require_login(request)
    if not user:
        return None, RedirectResponse(url="/login", status_code=302)
    if user.get("role") != "admin":
        flash(request, "Access denied", "danger")
        return None, RedirectResponse(url="/inventory", status_code=302)
    return user, None


def _normalize_mac(raw: str) -> str:
    clean = re.sub(r"[^a-fA-F0-9]", "", str(raw)).upper()
    if len(clean) != 12:
        raise ValueError(f"MAC address must have exactly 12 hex digits, got: {raw!r}")
    return ":".join(clean[i:i + 2] for i in range(0, 12, 2))


@router.get("/export")
def export_devices(request: Request):
    user, redirect = _require_admin(request)
    if redirect:
        return redirect

    conn = get_db_connection()
    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute(
            f"SELECT {', '.join(EXPORT_COLUMNS)} FROM devices ORDER BY id"
        )
        rows = cur.fetchall()
    finally:
        return_db_connection(conn)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Devices"
    ws.append(EXPORT_COLUMNS)
    for row in rows:
        ws.append([row.get(col) for col in EXPORT_COLUMNS])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type=XLSX_MIME,
        headers={"Content-Disposition": 'attachment; filename="devices_export.xlsx"'},
    )


@router.post("/import")
async def import_devices(request: Request, file: UploadFile = File(...)):
    user, redirect = _require_admin(request)
    if redirect:
        return redirect

    if not file.filename.lower().endswith(".xlsx"):
        return JSONResponse({"error": "Only .xlsx files are supported."}, status_code=400)

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        return JSONResponse({"error": "Could not parse file. Ensure it is a valid .xlsx file."}, status_code=400)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return JSONResponse({"error": "The file is empty."}, status_code=400)

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        return JSONResponse(
            {"error": f"Missing required columns: {', '.join(sorted(missing))}"},
            status_code=400,
        )

    col_index = {name: headers.index(name) for name in headers if name}

    inserted = 0
    overwritten = []
    errors = []

    conn = get_db_connection()
    try:
        cur = conn.cursor()
        for row_num, row in enumerate(rows[1:], start=2):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            def get(col):
                idx = col_index.get(col)
                val = row[idx] if idx is not None else None
                return str(val).strip() if val is not None else ""

            raw_mac = get("mac_address")
            try:
                mac = _normalize_mac(raw_mac)
            except ValueError as e:
                errors.append({"row": row_num, "message": str(e)})
                continue

            availability = get("availability")
            if availability not in VALID_AVAILABILITY:
                errors.append({
                    "row": row_num,
                    "message": f"Invalid availability: {availability!r}. Must be 'Available' or 'In Use'.",
                })
                continue

            device_model = get("device_model")
            if not device_model:
                errors.append({"row": row_num, "message": "device_model is required."})
                continue

            owner = get("owner") or None
            reporting_manager = get("reporting_manager") or None
            team = get("team") or None
            ip_address = get("ip_address") or None
            location = get("location") or None
            lease = get("lease") or None
            project = get("project") or None
            console = get("console") or None
            power = get("power") or None

            cur.execute("SELECT id FROM devices WHERE mac_address = %s", (mac,))
            existing = cur.fetchone()

            if existing:
                cur.execute(
                    """
                    UPDATE devices
                    SET device_model = %s, owner = %s, availability = %s,
                        reporting_manager = %s, team = %s, ip_address = %s,
                        location = %s, lease = %s, project = %s, console = %s, power = %s
                    WHERE mac_address = %s
                    """,
                    (device_model, owner, availability, reporting_manager,
                     team, ip_address, location, lease, project, console, power, mac),
                )
                overwritten.append(mac)
            else:
                cur.execute(
                    """
                    INSERT INTO devices
                        (mac_address, device_model, owner, availability,
                         reporting_manager, team, ip_address, location, lease,
                         project, console, power)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (mac, device_model, owner, availability, reporting_manager,
                     team, ip_address, location, lease, project, console, power),
                )
                inserted += 1

        conn.commit()
    except Exception as e:
        conn.rollback()
        return_db_connection(conn)
        return JSONResponse({"error": f"Database error: {e}"}, status_code=500)
    finally:
        return_db_connection(conn)

    return JSONResponse({
        "inserted": inserted,
        "overwritten": overwritten,
        "errors": errors,
    })
